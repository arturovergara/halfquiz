# Django Imports
from django import forms
from django.core.exceptions import ValidationError
from django.forms.models import inlineformset_factory

# 3rd Party Libraries
from openpyxl import load_workbook

from .fields import ExcelField
from .models import Game, GameQuestion, Option, Question, Topic


class TopicForm(forms.ModelForm):
    class Meta:
        model = Topic
        fields = ("name", "description")


class QuestionBulkCreateForm(forms.Form):
    topic = forms.ModelChoiceField(queryset=Topic.objects.all())
    questions_file = ExcelField()

    def clean(self):
        cleaned_data = super(QuestionBulkCreateForm, self).clean()
        wb = load_workbook(cleaned_data["questions_file"])
        sheet = wb.worksheets[0]
        questions = []

        for row in sheet.iter_rows(min_row=2, max_col=7, values_only=True):
            if (row[0] is None) and (row[1] is None):
                continue

            question_statement = str(row[0])
            question_time = str(row[1])
            correct_option = int(str(row[6]))
            options = [
                {"text": str(option), "is_right": i == correct_option}
                for i, option in enumerate(row[2:6], start=1)
            ]

            questions.append(
                {
                    "statement": question_statement,
                    "time": int(question_time),
                    "options": options,
                }
            )

        cleaned_data["questions"] = questions

        return cleaned_data

    def save(self):
        topic = self.cleaned_data["topic"]

        for question_data in self.cleaned_data["questions"]:
            question = Question.objects.create(
                statement=question_data["statement"],
                time=question_data["time"],
                topic=topic,
            )
            options = [
                Option(
                    text=option["text"],
                    is_right=option["is_right"],
                    question=question,
                )
                for option in question_data["options"]
            ]

            Option.objects.bulk_create(options)


class QuestionForm(forms.ModelForm):
    class Meta:
        model = Question
        fields = ("statement", "topic", "time")


class OptionForm(forms.ModelForm):
    class Meta:
        model = Option
        fields = ("text", "is_right")


class BaseOptionFormSet(forms.BaseInlineFormSet):
    def clean(self):
        if any(self.errors):
            return

        right_answers = []

        for form in self.forms:
            if self.can_delete and self._should_delete_form(form):
                continue

            is_right = form.cleaned_data.get("is_right")
            right_answers.append(is_right)

        if not any(right_answers):
            raise ValidationError("There must be at least one correct option")


OptionFormSet = inlineformset_factory(
    Question,
    Option,
    form=OptionForm,
    formset=BaseOptionFormSet,
    extra=4,
    max_num=4,
    min_num=2,
    validate_min=True,
)


class GameCreateForm(forms.Form):
    topic = forms.ModelChoiceField(queryset=Topic.objects.all())
    number_of_questions = forms.IntegerField(min_value=1)

    def create_game(self):
        topic = self.cleaned_data["topic"]
        number_of_questions = self.cleaned_data["number_of_questions"]
        game = Game.objects.create_random_game_by_topic(
            topic=topic, number_of_questions=number_of_questions
        )

        return game


class InGameQuestionForm(forms.ModelForm):
    class Meta:
        model = GameQuestion
        fields = ("answer",)
        widgets = {
            "answer": forms.RadioSelect,
        }

    def __init__(self, *args, **kwargs):
        super(InGameQuestionForm, self).__init__(*args, **kwargs)

        self.fields["answer"].empty_label = None
        self.fields["answer"].queryset = Option.objects.filter(
            question=self.instance.question
        )

    def save(self, commit=True):
        game_question = super(InGameQuestionForm, self).save(commit)
        game_question.game.answer_question()

        return game_question.game
